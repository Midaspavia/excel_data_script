import os
import pandas as pd
from excel_kennzahlen import fetch_excel_kennzahlen_by_ric, fetch_excel_kennzahlen_by_ric_filtered, clear_excel_cache
from refinitiv_integration import get_refinitiv_kennzahlen_for_companies, get_all_sector_averages
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import warnings

# KORRIGIERT: Unterdrücke openpyxl Warnungen über Datums-Formatierung
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

DATA_DIR = "excel_data/data"

def clean_refinitiv_field_name(field_name):
    """
    Entfernt TR. aus Refinitiv-Feldnamen, behält aber Period-Information bei
    Beispiel: TR.EBIT(Period=FY-1) → EBIT(Period=FY-1)
    """
    if field_name.startswith('TR.'):
        return field_name.replace('TR.', '')
    return field_name

def cleanup_temp_files():
    """Bereinigt temporäre Excel-Dateien (~$*.xlsx) nach der Ausführung"""
    print("🧹 BEREINIGE TEMPORÄRE DATEIEN...")

    directories = ["excel_data/", "excel_data/data/", "."]
    deleted_count = 0

    for directory in directories:
        if os.path.exists(directory):
            temp_files = glob.glob(os.path.join(directory, "~$*.xlsx"))
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                    deleted_count += 1
                except Exception as e:
                    pass

    if deleted_count > 0:
        print(f"✅ {deleted_count} temporäre Dateien bereinigt")

def process_companies():
    """Hauptfunktion zur Verarbeitung der Unternehmen"""
    start_time = time.time()
    print("🚀 STARTE OPTIMIERTE VERARBEITUNG...")

    try:
        # 1. Lese input_user.xlsx (SCHNELL)
        print("📖 Lese input_user.xlsx...")
        df_input = pd.read_excel("excel_data/input_user.xlsx")

        # Kennzahlen aus der ersten Zeile
        first_row = df_input.iloc[0]
        excel_fields = list(dict.fromkeys(df_input["Kennzahlen aus Excel"].dropna().astype(str).str.strip().tolist()))
        refinitiv_fields = list(dict.fromkeys(df_input["Kennzahlen aus Refinitiv"].dropna().astype(str).str.strip().tolist()))

        # Filter-Einstellungen
        sub_industry_filter = str(first_row.get("Sub-Industry", "")).strip().upper()
        focus_filter = str(first_row.get("Focus", "")).strip().upper()

        if focus_filter == "X":
            is_focus = True
            filter_type = "Focus"
        elif sub_industry_filter == "X":
            is_focus = False
            filter_type = "Sub-Industry"
        else:
            is_focus = False
            filter_type = "Sub-Industry (Default)"

        print(f"🎯 Filter: {filter_type}")
        print(f"📋 Excel-Kennzahlen: {len(excel_fields)}")
        print(f"📊 Refinitiv-Kennzahlen: {len(refinitiv_fields)}")

        # 2. SAMMLE ALLE INPUT-UNTERNEHMEN (SCHNELL)
        input_companies = []
        all_gics_sectors = set()

        for index, row in df_input.iterrows():
            input_name = str(row.iloc[0] if len(row) > 0 else "").strip()
            input_ric = str(row.iloc[1] if len(row) > 1 else "").strip()

            gics_sector = ""
            if "GICS Sector" in df_input.columns:
                gics_sector = str(row.get("GICS Sector", "")).strip()

            # Überspringe leere Zeilen
            if not input_name and not input_ric:
                continue
            if input_name.lower() in ["", "nan", "none"] and input_ric.lower() in ["", "nan", "none"]:
                continue

            if gics_sector and gics_sector.lower() not in ["", "nan", "none"]:
                all_gics_sectors.add(gics_sector)

            input_companies.append({
                'name': input_name if input_name.lower() not in ["", "nan", "none"] else None,
                'ric': input_ric if input_ric.lower() not in ["", "nan", "none"] else None,
                'gics_sector': gics_sector if gics_sector.lower() not in ["", "nan", "none"] else None,
                'row_number': index + 1
            })

        print(f"📋 {len(input_companies)} Input-Unternehmen")
        print(f"🏭 GICS Sektoren: {sorted(all_gics_sectors)}")

        # 3. OPTIMIERTE PEER-GROUP-VERARBEITUNG
        all_results = []
        processed_groups = set()

        for i, input_company in enumerate(input_companies, 1):
            print(f"\n🔍 {i}/{len(input_companies)}: Zeile {input_company['row_number']}")

            # Bestimme Suchstrategie
            start_company = None
            if input_company['ric']:
                print(f"   🎯 RIC: {input_company['ric']}")
                start_company = find_company_by_ric(input_company['ric'])
            elif input_company['name']:
                if len(input_company['name']) >= 4:
                    print(f"   🎯 Name: {input_company['name']}")
                    start_company = find_company_by_name(input_company['name'])
                else:
                    print(f"   ❌ Name zu kurz: {input_company['name']}")
                    continue
            else:
                print("   ❌ Weder RIC noch Name")
                continue

            if not start_company:
                print(f"   ❌ Nicht gefunden!")
                continue

            print(f"   ✅ {start_company['Name']} ({start_company['RIC']})")

            # KORRIGIERT: Bestimme Filter-Typ für JEDE ZEILE INDIVIDUELL
            current_row = df_input.iloc[input_company['row_number'] - 1]  # -1 wegen 0-basiertem Index

            row_sub_industry_filter = str(current_row.get("Sub-Industry", "")).strip().upper()
            row_focus_filter = str(current_row.get("Focus", "")).strip().upper()

            # Entscheide für diese spezifische Zeile
            if row_focus_filter == "X":
                use_focus_for_this_row = True
                current_filter_type = "Focus"
            elif row_sub_industry_filter == "X":
                use_focus_for_this_row = False
                current_filter_type = "Sub-Industry"
            else:
                # Fallback: Verwende den globalen Filter
                use_focus_for_this_row = is_focus
                current_filter_type = filter_type

            print(f"   🎯 Filter für diese Zeile: {current_filter_type}")

            # KORRIGIERT: Erstelle eindeutigen Gruppenschlüssel der Kollisionen verhindert
            if use_focus_for_this_row:
                group_key = f"Focus_{start_company.get('Focus', 'Unknown')}"
                peer_group_type = "Focus"
            else:
                group_key = f"SubIndustry_{start_company.get('Sub-Industry', 'Unknown')}"
                peer_group_type = "Sub-Industry"

            # WICHTIG: Überspringe nur wenn die GLEICHE Gruppe bereits verarbeitet wurde
            # Aber erlaube verschiedene Gruppen-Typen
            if group_key in processed_groups:
                print(f"   ⏭️  Peer-Gruppe '{group_key}' bereits verarbeitet - überspringe")
                continue

            print(f"   🆕 Neue Peer-Gruppe wird verarbeitet: {group_key}")
            processed_groups.add(group_key)

            # 4. FINDE PEER-GRUPPE (KORRIGIERT)
            print(f"   🔍 Suche {peer_group_type}-Peer-Gruppe...")

            peer_companies = []
            if use_focus_for_this_row and start_company.get('Focus'):
                focus_value = start_company['Focus']
                print(f"     🎯 Focus-Suche: '{focus_value}'")
                peer_companies = find_companies_by_focus(focus_value)
            elif start_company.get('Sub-Industry'):
                sub_industry_value = start_company['Sub-Industry']
                print(f"     🏭 Sub-Industry-Suche: '{sub_industry_value}'")
                peer_companies = find_companies_by_sub_industry(sub_industry_value)

            if not peer_companies:
                print(f"     ⚠️ Keine Peer-Gruppe gefunden, verarbeite nur das Unternehmen")
                peer_companies = [start_company]

            print(f"     ✅ {peer_group_type}-Peer-Gruppe: {len(peer_companies)} Unternehmen")

            # 5. VERARBEITE PEER-GRUPPE MIT OPTIMIERTER EXCEL-KENNZAHLEN-LOGIK
            peer_results = []

            for j, company in enumerate(peer_companies, 1):
                print(f"     🏢 {j}/{len(peer_companies)}: {company['Name']}")

                # Excel-Kennzahlen (BEIBEHALTENE OPTIMIERTE LOGIK)
                excel_data = {}
                if excel_fields:
                    # Verwende GICS Sector-Filter falls verfügbar
                    gics_filter = [input_company['gics_sector']] if input_company.get('gics_sector') else None
                    excel_data = fetch_excel_kennzahlen_by_ric_filtered(
                        company['RIC'],
                        excel_fields,
                        gics_sectors=gics_filter
                    )

                # Refinitiv-Kennzahlen
                refinitiv_data = {}
                if refinitiv_fields:
                    # KORRIGIERT: Übergebe Company-Dictionary statt nur RIC-String
                    company_list = [{'RIC': company['RIC']}]
                    refinitiv_result = get_refinitiv_kennzahlen_for_companies(company_list, refinitiv_fields)
                    if company['RIC'] in refinitiv_result:
                        refinitiv_data = refinitiv_result[company['RIC']]

                # Bestimme GICS Sektor für das Unternehmen
                gics_sector = determine_gics_sector(company['RIC'])

                # Kombiniere Ergebnisse
                result_row = {
                    'Name': company['Name'],
                    'RIC': company['RIC'],
                    'GICS Sector': gics_sector,  # GICS Sektor hinzufügen
                    'Sub-Industry': company.get('Sub-Industry', ''),
                    'Focus': company.get('Focus', ''),
                    'Peer_Group_Type': peer_group_type,  # Neu: Kennzeichnung der Gruppe
                    'Input_Row': f"Zeile {input_company['row_number']}" if peer_group_type in ['Focus', 'Sub-Industry'] else '',  # Zeigt Input-Zeile
                }

                # Füge Excel-Kennzahlen hinzu
                for field in excel_fields:
                    result_row[field] = excel_data.get(field, '')

                # Füge Refinitiv-Kennzahlen hinzu
                for field in refinitiv_fields:
                    clean_field = clean_refinitiv_field_name(field)
                    # KORRIGIERT: Verwende den ursprünglichen Feldnamen für den Lookup
                    original_field_value = refinitiv_data.get(field, '')
                    clean_field_value = refinitiv_data.get(clean_field, '')
                    # Nimm den Wert, der nicht leer ist
                    final_value = original_field_value if original_field_value else clean_field_value
                    result_row[clean_field] = final_value

                peer_results.append(result_row)
                print(f"       ✅ {len(excel_data)} Excel + {len(refinitiv_data)} Refinitiv")

            # Füge alle Peer-Ergebnisse zur Gesamt-Liste hinzu
            all_results.extend(peer_results)
            print(f"   📊 {peer_group_type}-Peer-Gruppe verarbeitet: {len(peer_results)} Unternehmen hinzugefügt")

        # 6. Speichere Output mit schönem Design
        if all_results:
            output_path = "excel_data/output.xlsx"
            df_output = pd.DataFrame(all_results)

            print(f"\n📊 INSGESAMT {len(all_results)} UNTERNEHMEN VERARBEITET")
            print("💾 Speichere in output.xlsx...")

            # KORRIGIERT: Stelle sicher, dass Output-Verzeichnis existiert
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                print(f"📁 Erstelle fehlendes Verzeichnis: {output_dir}")
                os.makedirs(output_dir, exist_ok=True)

            # 🔢 BERECHNE DURCHSCHNITTE FÜR EXCEL-KENNZAHLEN
            print("\n🔢 BERECHNE DURCHSCHNITTE FÜR EXCEL-KENNZAHLEN...")
            df_output_with_averages = calculate_excel_averages(df_output, excel_fields)

            # 🔢 BERECHNE REFINITIV-DURCHSCHNITTE NACH SEKTOR
            print("\n🔢 BERECHNE REFINITIV-DURCHSCHNITTE NACH SEKTOR...")
            if refinitiv_fields:
                df_output_with_averages = calculate_refinitiv_averages_by_sector(df_output_with_averages, refinitiv_fields)

            # KORRIGIERT: Filtere Output-DataFrame, um nur angeforderte Kennzahlen zu behalten (WIE IN DER FUNKTIONIERENDEN VERSION)
            print(f"\n🔍 FILTERE OUTPUT AUF NUR ANGEFORDERTE KENNZAHLEN...")

            # Basis-Spalten die immer beibehalten werden (WIE IN DER FUNKTIONIERENDEN VERSION)
            base_columns = ['Name', 'RIC', 'GICS Sector', 'Sub-Industry', 'Focus', 'Peer_Group_Type', 'Input_Row']

            # Sammle alle erlaubten Spalten
            allowed_columns = base_columns.copy()
            allowed_columns.extend(excel_fields)  # Angeforderte Excel-Kennzahlen

            # Füge Refinitiv-Kennzahlen hinzu (mit und ohne TR. Präfix)
            for ref_field in refinitiv_fields:
                allowed_columns.append(ref_field)  # Original (z.B. TR.EBIT)
                clean_field = clean_refinitiv_field_name(ref_field)
                allowed_columns.append(clean_field)  # Ohne TR. (z.B. EBIT)

            # Filtere DataFrame auf nur erlaubte Spalten
            existing_allowed_columns = [col for col in allowed_columns if col in df_output_with_averages.columns]
            df_output_cleaned = df_output_with_averages[existing_allowed_columns].copy()

            # KORRIGIERT: Entferne leere Spalten (Spalten mit leerem Namen oder nur leeren Werten)
            columns_to_keep = []
            for col in df_output_cleaned.columns:
                # Überspringe Spalten mit leerem Namen
                if col == '' or str(col).strip() == '':
                    continue
                # Überspringe Spalten die nur leere Werte enthalten
                if df_output_cleaned[col].isna().all() or (df_output_cleaned[col].astype(str).str.strip() == '').all():
                    continue
                columns_to_keep.append(col)

            # Filtere DataFrame auf nur sinnvolle Spalten
            df_output_cleaned = df_output_cleaned[columns_to_keep].copy()

            print(f"   📊 Ursprüngliche Spalten: {len(df_output_with_averages.columns)}")
            print(f"   ✅ Gefilterte Spalten: {len(df_output_cleaned.columns)}")
            print(f"   🧹 Leere Spalten entfernt: {len(existing_allowed_columns) - len(columns_to_keep)}")
            print(f"   📋 Behaltene Spalten: {list(df_output_cleaned.columns)}")

            # Erstelle schön formatierte Excel-Datei (WIE IN DER FUNKTIONIERENDEN VERSION)
            create_beautiful_excel_output(df_output_cleaned, output_path, excel_fields, len(all_results))

            print(f"\n✅ SCHÖN FORMATIERTES OUTPUT GESPEICHERT: {output_path}")
            print(f"📊 {len(all_results)} Unternehmen + {len(df_output_cleaned) - len(all_results)} Durchschnittswerte = {len(df_output_cleaned)} Zeilen insgesamt mit {len(df_output_cleaned.columns)} Spalten")

            # Zeige Übersicht (WIE IN DER FUNKTIONIERENDEN VERSION)
            print(f"\n📋 ERGEBNIS-ÜBERSICHT:")
            for i, result in enumerate(all_results, 1):
                print(f"\n{i}. {result['Name']} ({result['RIC']}) - {result.get('Input_Row', '')}")
                print(f"   GICS Sector: {result.get('GICS Sector', 'N/A')}")
                print(f"   Sub-Industry: {result.get('Sub-Industry', 'N/A')}")
                print(f"   Focus: {result.get('Focus', 'N/A')}")

                # Zeige alle Excel-Kennzahlen
                for field in excel_fields:
                    value = result.get(field, 'N/A')
                    if value != 'N/A' and pd.notna(value):
                        print(f"   [Excel] {field}: {value}")
                    else:
                        print(f"   [Excel] {field}: ❌ Nicht gefunden")

                # Sammle alle Refinitiv-relevanten Spalten aus dem tatsächlichen DataFrame
                actual_refinitiv_columns = []

                # 1. Alle ursprünglich angeforderten Refinitiv-Felder
                for field in refinitiv_fields:
                    actual_refinitiv_columns.append(field)

                # 2. Alle Spalten im result, die wie Refinitiv-Felder aussehen
                for key in result.keys():
                    # Überspringt Basis-Spalten und Excel-Kennzahlen
                    if key not in ['Name', 'RIC', 'GICS Sector', 'Sub-Industry', 'Focus', 'Peer_Group_Type', 'Input_Row'] and key not in excel_fields:
                        # Prüft, ob es ein potentielles Refinitiv-Feld ist
                        if (key.startswith('TR.') or
                            any(key.upper() == ref_field.replace('TR.', '').upper() for ref_field in refinitiv_fields)):
                            if key not in actual_refinitiv_columns:
                                actual_refinitiv_columns.append(key)

                # Entferne Duplikate und behalte Reihenfolge
                actual_refinitiv_columns = list(dict.fromkeys(actual_refinitiv_columns))

                # Zeige alle gefundenen Refinitiv-Kennzahlen
                for field in actual_refinitiv_columns:
                    # Suche nach der Spalte im Result
                    found_value = None
                    found_key = None

                    # Direkte Suche nach dem Feld
                    if field in result:
                        found_value = result[field]
                        found_key = field
                    else:
                        # Erweiterte Suche für ursprünglich angeforderten Felder
                        cleaned_field = field.replace("TR.", "") if field.startswith("TR.") else field
                        if cleaned_field in result:
                            found_value = result[cleaned_field]
                            found_key = cleaned_field
                        else:
                            # Fuzzy-Suche nach ähnlichen Feldern
                            for key, value in result.items():
                                if (field.lower() in key.lower() or
                                    cleaned_field.lower() in key.lower() or
                                    key.lower() in field.lower()):
                                    found_value = value
                                    found_key = key
                                    break

                    if found_value is not None and pd.notna(found_value) and str(found_value).strip() != '':
                        # Bestimme Label für Ausgabe
                        if field in refinitiv_fields:
                            display_label = f"[Refinitiv] {field}"
                        else:
                            display_label = f"[Refinitiv*] {field}"  # * für neu erstellte Spalten

                        if found_key != field:
                            print(f"   {display_label} (als '{found_key}'): {found_value}")
                        else:
                            print(f"   {display_label}: {found_value}")
                    else:
                        print(f"   [Refinitiv] {field}: ❌ Nicht gefunden")

            # Zeige GICS-Sektor-Durchschnitte für Refinitiv-Kennzahlen (WIE IN DER FUNKTIONIERENDEN VERSION)
            if refinitiv_fields:
                # Sammle alle Sektor-Durchschnitte aus dem DataFrame
                sector_avg_rows = df_output_cleaned[df_output_cleaned['Name'].str.contains('🏭 Ø', na=False)]

                if not sector_avg_rows.empty:
                    print(f"\n🏭 GICS-SEKTOR-DURCHSCHNITTE (REFINITIV):")
                    for _, sector_row in sector_avg_rows.iterrows():
                        sector_name = sector_row['Name'].replace('🏭 Ø ', '')
                        print(f"\n📊 {sector_name}:")
                        for field in refinitiv_fields:
                            clean_field = clean_refinitiv_field_name(field)
                            # Suche nach dem Wert in der Sektor-Zeile
                            value = None
                            for possible_key in [field, clean_field, field.replace('TR.', ''), clean_field]:
                                try:
                                    # KORRIGIERT: Verwende hasattr und direkten Zugriff auf Series
                                    if hasattr(sector_row, possible_key) and pd.notna(getattr(sector_row, possible_key, None)):
                                        potential_value = getattr(sector_row, possible_key)
                                        if str(potential_value).strip() != '':
                                            value = potential_value
                                            break
                                    # Alternative: Verwende Dictionary-ähnlichen Zugriff
                                    elif possible_key in sector_row and pd.notna(sector_row[possible_key]):
                                        potential_value = sector_row[possible_key]
                                        if str(potential_value).strip() != '':
                                            value = potential_value
                                            break
                                except (KeyError, ValueError, AttributeError):
                                    continue

                            if value is not None:
                                print(f"   📈 {field}: {value:,.4f} (Sektor-Durchschnitt)")
                            else:
                                print(f"   📈 {field}: ❌ Nicht verfügbar")

            end_time = time.time()
            print(f"\n🎉 PEER-GROUP-ANALYSE ERFOLGREICH! Ausführungszeit: {end_time - start_time:.1f}s")
            print(f"📊 Ergebnisse: {len(all_results)} Unternehmen in verschiedenen Peer-Gruppen")

            # Zeige Zusammenfassung der Peer-Gruppen
            focus_companies = [r for r in all_results if r.get('Peer_Group_Type') == 'Focus']
            sub_industry_companies = [r for r in all_results if r.get('Peer_Group_Type') == 'Sub-Industry']

            print(f"\n📋 PEER-GRUPPEN ÜBERSICHT:")
            if focus_companies:
                print(f"   🎯 Focus-Gruppen: {len(focus_companies)} Unternehmen")
            if sub_industry_companies:
                print(f"   🏭 Sub-Industry-Gruppen: {len(sub_industry_companies)} Unternehmen")

            print(f"\n✅ OUTPUT ERFOLGREICH GESPEICHERT: {output_path}")

        else:
            print("❌ Keine Ergebnisse zum Schreiben")

        # Bereinige temporäre Dateien
        cleanup_temp_files()

        return all_results

    except Exception as e:
        print(f"\n❌ FEHLER: {e}")
        import traceback
        traceback.print_exc()
        cleanup_temp_files()
        return []
def find_company_by_ric(ric):
    """Finde Unternehmen anhand des RIC - direkte Index-Positionen"""
    print(f"🔍 RIC-Suche: '{ric}' (RIC=Spalte E, Focus=Spalte D, Sub-Industry=Spalte C)")

    for file in os.listdir(DATA_DIR):
        if not file.endswith(".xlsx") or file.startswith("~$"):
            continue

        file_path = os.path.join(DATA_DIR, file)

        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                # Nur relevante Sheets
                if not any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                    continue

                try:
                    # Lese mit Header=2 (Zeile 3)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=200)

                    # Prüfe ob genügend Spalten vorhanden sind
                    if len(df.columns) < 5:
                        continue

                    # RIC ist in Spalte E (Index 4)
                    ric_col = df.columns[4]

                    # Suche nach dem spezifischen RIC
                    matches = df[df[ric_col].astype(str).str.upper().str.strip() == ric.upper().strip()]

                    if not matches.empty:
                        row = matches.iloc[0]

                        # Direkte Index-Zugriffe
                        sub_industry = str(row.iloc[2]).strip()  # Spalte C
                        focus_value = str(row.iloc[3]).strip()   # Spalte D
                        ric_value = str(row.iloc[4]).strip()     # Spalte E

                        # KORRIGIERT: Robuste Name-Extraktion aus Spalte A (Holding) oder B (Universe)
                        name_value = "Unknown"

                        # Versuche zuerst Spalte A (Holding)
                        if len(df.columns) > 0:
                            holding_name = str(row.iloc[0]).strip()
                            if holding_name and holding_name.lower() not in ['nan', 'none', ''] and len(holding_name) > 2:
                                name_value = holding_name

                        # Falls Spalte A leer oder ungültig, versuche Spalte B (Universe)
                        if name_value == "Unknown" and len(df.columns) > 1:
                            universe_name = str(row.iloc[1]).strip()
                            if universe_name and universe_name.lower() not in ['nan', 'none', ''] and len(universe_name) > 2:
                                name_value = universe_name

                        # Falls beide Spalten leer, verwende generischen Fallback-Namen
                        if name_value == "Unknown":
                            name_value = f"Company_{ric_value}"

                        company = {
                            "Name": name_value,
                            "RIC": ric_value,
                            "Sub-Industry": sub_industry,
                            "Focus": focus_value
                        }
                        print(f"✅ GEFUNDEN: {company['Name']} ({company['RIC']})")
                        print(f"   Sub-Industry (Spalte C): '{company['Sub-Industry']}'")
                        print(f"   Focus (Spalte D): '{company['Focus']}'")
                        return company

                except Exception as e:
                    continue

        except Exception as e:
            continue

    print(f"❌ RIC '{ric}' nicht gefunden")
    return None


def find_companies_by_focus(focus):
    """Suche alle Unternehmen mit gleichem Focus (Spalte D)"""
    companies = []
    print(f"🔍 Focus-Suche in Spalte D: '{focus}'")

    for file in os.listdir(DATA_DIR):
        if not file.endswith(".xlsx") or file.startswith("~$"):
            continue

        file_path = os.path.join(DATA_DIR, file)
        print(f"📁 Durchsuche: {file}")

        try:
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                # Nur relevante Sheets
                if not any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                    continue

                try:
                    # Lese mit Header=2 (Zeile 3)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)

                    # Prüfe ob genügend Spalten vorhanden sind
                    if len(df.columns) < 5:
                        continue

                    # Suche in Spalte D (Index 3) nach gleichem Focus
                    found_in_sheet = 0
                    for _, row in df.iterrows():
                        if len(row) >= 5 and pd.notna(row.iloc[1]) and pd.notna(row.iloc[3]) and pd.notna(row.iloc[4]):
                            row_focus = str(row.iloc[3]).strip()  # Spalte D

                            if row_focus == focus.strip():
                                company = {
                                    "Name": str(row.iloc[1]).strip(),       # Spalte B
                                    "RIC": str(row.iloc[4]).strip(),        # Spalte E
                                    "Sub-Industry": str(row.iloc[2]).strip(), # Spalte C
                                    "Focus": row_focus                       # Spalte D
                                }

                                # Vermeide Duplikate
                                if not any(c["RIC"] == company["RIC"] for c in companies):
                                    companies.append(company)
                                    found_in_sheet += 1

                    if found_in_sheet > 0:
                        print(f"  📄 {sheet_name}: {found_in_sheet} Unternehmen gefunden")

                except Exception as e:
                    continue

        except Exception as e:
            continue

    print(f"📊 GESAMT: {len(companies)} Unternehmen mit Focus '{focus}' gefunden")
    return companies


def find_companies_by_sub_industry(sub_industry):
    """Suche alle Unternehmen mit gleicher Sub-Industry (Spalte C)"""
    companies = []
    print(f"🔍 Sub-Industry-Suche in Spalte C: '{sub_industry}'")

    for file in os.listdir(DATA_DIR):
        if not file.endswith(".xlsx") or file.startswith("~$"):
            continue

        file_path = os.path.join(DATA_DIR, file)
        print(f"📁 Durchsuche: {file}")

        try:
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                # Nur relevante Sheets
                if not any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                    continue

                try:
                    # Lese mit Header=2 (Zeile 3)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)

                    # Prüfe ob genügend Spalten vorhanden sind
                    if len(df.columns) < 5:
                        continue

                    # Suche in Spalte C (Index 2) nach gleicher Sub-Industry
                    found_in_sheet = 0
                    for _, row in df.iterrows():
                        if len(row) >= 5 and pd.notna(row.iloc[1]) and pd.notna(row.iloc[2]) and pd.notna(row.iloc[4]):
                            row_sub_industry = str(row.iloc[2]).strip()  # Spalte C

                            if row_sub_industry == sub_industry.strip():
                                company = {
                                    "Name": str(row.iloc[1]).strip(),       # Spalte B
                                    "RIC": str(row.iloc[4]).strip(),        # Spalte E
                                    "Sub-Industry": row_sub_industry,        # Spalte C
                                    "Focus": str(row.iloc[3]).strip()       # Spalte D
                                }

                                # Vermeide Duplikate
                                if not any(c["RIC"] == company["RIC"] for c in companies):
                                    companies.append(company)
                                    found_in_sheet += 1

                    if found_in_sheet > 0:
                        print(f"  📄 {sheet_name}: {found_in_sheet} Unternehmen gefunden")

                except Exception as e:
                    continue

        except Exception as e:
            continue

    print(f"📊 GESAMT: {len(companies)} Unternehmen mit Sub-Industry '{sub_industry}' gefunden")
    return companies


def get_kennzahlen_for_company(ric, fields):
    """Sammelt alle gewünschten Kennzahlen für ein Unternehmen basierend auf RIC (nutzt robusten Import aus excel_kennzahlen.py)"""
    return fetch_excel_kennzahlen_by_ric(ric, fields)

def find_company_by_name(name):
    """Finde Unternehmen anhand des Namens - Suche in Holding/Universe"""
    print(f"🔍 Name-Suche: '{name}' (Teilwort-Suche in Holding/Universe)")

    # Prüfe 4-Zeichen-Regel
    if len(name) < 4:
        print(f"❌ Name '{name}' zu kurz (mindestens 4 Zeichen erforderlich)")
        return None

    for file in os.listdir(DATA_DIR):
        if not file.endswith(".xlsx") or file.startswith("~$"):
            continue

        file_path = os.path.join(DATA_DIR, file)

        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                # Nur relevante Sheets
                if not any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                    continue

                try:
                    # Lese mit Header=2 (Zeile 3)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)

                    # Prüfe ob genügend Spalten vorhanden sind
                    if len(df.columns) < 5:
                        continue

                    # KORRIGIERT: Suche sowohl in Spalte A (Holding) als auch Spalte B (Universe)
                    holding_col = df.columns[0]  # Spalte A (Holding)
                    universe_col = df.columns[1]  # Spalte B (Universe)

                    # Suche in beiden Spalten
                    holding_matches = df[df[holding_col].astype(str).str.contains(name, case=False, na=False)]
                    universe_matches = df[df[universe_col].astype(str).str.contains(name, case=False, na=False)]

                    # Kombiniere beide Ergebnisse, bevorzuge aber Holding-Treffer
                    matches = holding_matches if not holding_matches.empty else universe_matches

                    if not matches.empty:
                        row = matches.iloc[0]

                        # Direkte Index-Zugriffe
                        holding_value = str(row.iloc[0]).strip()   # Spalte A (Holding)
                        universe_value = str(row.iloc[1]).strip()  # Spalte B (Universe)
                        sub_industry = str(row.iloc[2]).strip()    # Spalte C
                        focus_value = str(row.iloc[3]).strip()     # Spalte D
                        ric_value = str(row.iloc[4]).strip()       # Spalte E

                        # Bestimme den Namen (Holding hat Priorität)
                        if holding_value and holding_value != 'nan' and len(holding_value.strip()) > 2:
                            name_value = holding_value
                            found_in = "Holding"
                        else:
                            name_value = universe_value
                            found_in = "Universe"

                        company = {
                            "Name": name_value,
                            "RIC": ric_value,
                            "Sub-Industry": sub_industry,
                            "Focus": focus_value
                        }
                        print(f"✅ GEFUNDEN: {company['Name']} ({company['RIC']}) in {found_in}-Spalte")
                        print(f"   Sub-Industry (Spalte C): '{company['Sub-Industry']}'")
                        print(f"   Focus (Spalte D): '{company['Focus']}'")
                        return company

                except Exception as e:
                    continue

        except Exception as e:
            continue

    print(f"❌ Name '{name}' nicht gefunden")
    return None

def create_beautiful_excel_output(df, output_path, excel_fields, actual_company_count=None):
    """Erstellt eine wunderschön formatierte Excel-Datei mit professionellem Design"""
    print("🎨 ERSTELLE SCHÖNES EXCEL-DESIGN...")

    # Speichere DataFrame als Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Financial Analysis', index=False)

    # Lade Workbook für Formatierung
    wb = load_workbook(output_path)
    ws = wb['Financial Analysis']

    # 🎨 FARB-SCHEMA (NUR ALTERNIERENDE FARBEN)
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")  # Dunkles Blau für Header
    alternating_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")  # Sehr helles Grau
    white_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")  # Weiß

    # 📝 SCHRIFT-STILE
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    company_font = Font(name="Calibri", size=11, bold=True, color="1f4e79")
    data_font = Font(name="Calibri", size=10, color="2f2f2f")

    # 📐 ALIGNMENT
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center")

    # 🔳 BORDERS
    thin_border = Border(
        left=Side(style="thin", color="b0b0b0"),
        right=Side(style="thin", color="b0b0b0"),
        top=Side(style="thin", color="b0b0b0"),
        bottom=Side(style="thin", color="b0b0b0")
    )

    thick_border = Border(
        left=Side(style="medium", color="1f4e79"),
        right=Side(style="medium", color="1f4e79"),
        top=Side(style="medium", color="1f4e79"),
        bottom=Side(style="medium", color="1f4e79")
    )

    # 1️⃣ HEADER-ZEILE FORMATIEREN
    print("  🎯 Formatiere Header...")

    # Berechne dynamische Header-Höhe basierend auf Zeilenumbrüchen
    max_lines = 1
    for col_num, cell in enumerate(ws[1], 1):
        col_name = df.columns[col_num - 1]

        # Zähle Zeilenumbrüche in Spalten-Namen
        line_count = col_name.count('\n') + 1
        max_lines = max(max_lines, line_count)

        # Formatiere Header-Zelle
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thick_border

    # Setze dynamische Header-Höhe (15 Pixel pro Zeile + 10 Pixel Padding)
    dynamic_header_height = max_lines * 15 + 10
    ws.row_dimensions[1].height = max(25, dynamic_header_height)  # Minimum 25 Pixel

    print(f"  📏 Header-Höhe: {ws.row_dimensions[1].height}px ({max_lines} Zeilen)")

    # 2️⃣ SPALTEN-KATEGORIEN BESTIMMEN
    company_cols = ['Name', 'RIC']  # Unternehmensdaten
    category_cols = ['Sub-Industry', 'Focus', 'GICS\nSektor']  # Kategorien
    metric_cols = [col for col in df.columns if col not in company_cols + category_cols + ['Peer_Group_Type', 'Input_Row', 'Sector']]  # Kennzahlen

    # 3️⃣ DATENZEILEN FORMATIEREN
    print("  🎯 Formatiere Datenzeilen...")
    for row_num in range(2, len(df) + 2):
        # Alternierend gefärbte Zeilen für bessere Lesbarkeit
        is_even = (row_num % 2 == 0)
        row_fill = alternating_fill if is_even else white_fill

        for col_num, cell in enumerate(ws[row_num], 1):
            col_name = df.columns[col_num - 1]

            # Basis-Formatierung
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill  # Nur alternierende Farben

            # Spezielle Formatierung je Spalten-Typ
            if col_name in company_cols:
                cell.font = company_font if col_name == 'Name' else Font(name="Calibri", size=10, bold=True, color="1f4e79")
                cell.alignment = left_alignment
            elif col_name in category_cols:
                cell.alignment = center_alignment
            elif col_name in metric_cols:
                cell.alignment = right_alignment

                # Formatiere Zahlen schön
                if cell.value and str(cell.value).replace('.', '').replace('-', '').isdigit():
                    try:
                        num_val = float(cell.value)
                        if abs(num_val) >= 1:
                            cell.number_format = '#,##0.00'  # Mit Tausender-Trennzeichen
                        else:
                            cell.number_format = '0.0000'    # Mehr Dezimalstellen für kleine Zahlen
                    except:
                        pass
            else:
                cell.alignment = left_alignment

        # Zeilenhöhe optimieren
        ws.row_dimensions[row_num].height = 20

    # 4️⃣ SPALTENBREITEN AUTOMATISCH ANPASSEN
    print("  🎯 Optimiere Spaltenbreiten...")
    for col_num, column in enumerate(ws.columns, 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        col_name = df.columns[col_num - 1]

        # Berechne optimale Breite basierend auf Inhalt
        max_length = 0
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        # Setze minimale und maximale Breiten je nach Spalten-Typ
        if col_name == 'Name':
            width = min(max(max_length + 2, 25), 40)  # Name: 25-40 Zeichen
        elif col_name == 'RIC':
            width = min(max(max_length + 2, 12), 15)  # RIC: 12-15 Zeichen
        elif col_name in category_cols:
            width = min(max(max_length + 2, 15), 25)  # Kategorien: 15-25 Zeichen
        elif col_name in metric_cols:
            width = min(max(max_length + 2, 12), 18)  # Kennzahlen: 12-18 Zeichen
        else:
            width = min(max(max_length + 2, 10), 30)  # Standard: 10-30 Zeichen

        ws.column_dimensions[col_letter].width = width

    # 5️⃣ CONDITIONAL FORMATTING ENTFERNT
    # (Keine Farbvergleiche mehr für Kennzahlen-Spalten)

    # 6️⃣ TITEL UND METADATA HINZUFÜGEN
    print("  🎯 Füge Titel hinzu...")
    # Neue Zeile oben einfügen für Titel
    ws.insert_rows(1)

    # Titel erstellen
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📊 FINANCIAL ANALYSIS REPORT - {df['Sub-Industry'].iloc[0] if len(df) > 0 else 'PEER ANALYSIS'}"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1f4e79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Titel über alle Spalten mergen
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Titel-Zeile höher machen
    ws.row_dimensions[1].height = 35

    # Titel-Hintergrund
    for col in range(1, len(df.columns) + 1):
        ws.cell(row=1, column=col).fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
        ws.cell(row=1, column=col).border = thick_border

    # KORRIGIERT: Header-Höhe nach Titel-Einfügung neu setzen (jetzt Zeile 2)
    ws.row_dimensions[2].height = max(25, dynamic_header_height)
    print(f"  📏 Header-Höhe korrigiert: {ws.row_dimensions[2].height}px ({max_lines} Zeilen) - Zeile 2")

    # 7️⃣ FREEZE PANES FÜR BESSERE NAVIGATION
    ws.freeze_panes = "A3"  # Freeze Header und Titel

    # 8️⃣ METADATA AM ENDE HINZUFÜGEN
    last_row = len(df) + 3
    metadata_cell = ws.cell(row=last_row, column=1)

    # KORRIGIERT: Verwende die tatsächliche Anzahl der Unternehmen (ohne Durchschnitte)
    company_count = actual_company_count if actual_company_count is not None else len(df)

    metadata_cell.value = f"📅 Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')} | 📊 Companies: {company_count}"
    metadata_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    metadata_cell.alignment = left_alignment

    # Metadata über mehrere spalten mergen
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=min(6, len(df.columns)))

    # Speichere formatierte Datei
    wb.save(output_path)

    print("  ✨ Excel-Formatierung abgeschlossen!")
    print(f"  📋 Titel: {title_cell.value}")
    print(f"  📊 {len(df)} Unternehmen formatiert")
    print(f"  📈 {len(metric_cols)} Kennzahlen mit Conditional Formatting")
    print(f"  💾 Datei gespeichert: {output_path}")

def calculate_excel_averages(df, excel_fields):
    """Berechnet die Durchschnitte für Excel-Kennzahlen nach Sub-Industry und Focus-Gruppen"""
    print("🔢 BERECHNE DURCHSCHNITTE FÜR EXCEL-KENNZAHLEN...")

    # Filtere nur die Spalten, die mit Excel-Kennzahlen gefüllt sind
    excel_columns = [field for field in excel_fields if field in df.columns]

    if not excel_columns:
        print("⚠️ Keine Excel-Kennzahlen gefunden, überspringe Durchschnittsberechnung")
        return df

    print(f"📊 Berechne Durchschnitte für: {excel_columns}")

    # Konvertiere Excel-Kennzahlen zu numerischen Werten
    df_numeric = df.copy()
    for col in excel_columns:
        df_numeric[col] = pd.to_numeric(df_numeric[col], errors='coerce')

    # 1. SUB-INDUSTRY DURCHSCHNITTE (ALLE UNTERNEHMEN AUS DEN EXCEL-DATEIEN)
    print("   🏭 Berechne Sub-Industry Durchschnitte (alle verfügbaren Unternehmen)...")

    # Hole alle eindeutigen Sub-Industries aus dem Output
    unique_sub_industries = df_numeric['Sub-Industry'].dropna().unique()

    for sub_industry in unique_sub_industries:
        if sub_industry and sub_industry.strip():
            print(f"     🔍 Suche alle Unternehmen der Sub-Industry: '{sub_industry}'")

            # Hole ALLE Unternehmen dieser Sub-Industry aus den Excel-Dateien
            all_companies_in_sub_industry = find_companies_by_sub_industry(sub_industry)

            if len(all_companies_in_sub_industry) > 1:
                # Sammle Excel-Kennzahlen für ALLE Unternehmen der Sub-Industry
                all_sub_industry_data = []
                print(f"       �� Verarbeite {len(all_companies_in_sub_industry)} Unternehmen...")

                for i, company in enumerate(all_companies_in_sub_industry, 1):
                    if i <= 5 or i % 20 == 0:  # Zeige nur jeden 20. nach den ersten 5
                        print(f"         {i}/{len(all_companies_in_sub_industry)}: {company['Name']}")

                    company_data = get_kennzahlen_for_company(company['RIC'], excel_columns)
                    if company_data:
                        # Füge Basis-Informationen hinzu
                        company_data.update({
                            'Name': company['Name'],
                            'RIC': company['RIC'],
                            'Sub-Industry': company.get('Sub-Industry', ''),
                            'Focus': company.get('Focus', '')
                        })
                        all_sub_industry_data.append(company_data)

                if all_sub_industry_data:
                    # Erstelle DataFrame für alle Sub-Industry Unternehmen
                    df_sub_industry = pd.DataFrame(all_sub_industry_data)

                    # Konvertiere zu numerischen Werten
                    for col in excel_columns:
                        df_sub_industry[col] = pd.to_numeric(df_sub_industry[col], errors='coerce')

                    # Berechne Durchschnitte
                    avg_row = {
                        'Name': f'💼 Ø {sub_industry}',
                        'RIC': '',
                        'GICS Sector': '',  # Hinzufügen der GICS Sector Spalte
                        'Sub-Industry': sub_industry,
                        'Focus': '',
                        'Peer_Group_Type': '',  # Leeres Peer_Group_Type für Sub-Industry-Durchschnitte
                        'Input_Row': '',  # Leeres Input_Row für Durchschnitte
                        'Input_Source': 'Durchschnitt (Branche)'
                    }

                    for col in excel_columns:
                        valid_values = df_sub_industry[col].dropna()
                        if len(valid_values) > 0:
                            avg_row[col] = valid_values.mean()
                            print(f"       📈 {col}: {avg_row[col]:.4f} (aus {len(valid_values)} von {len(df_sub_industry)} Unternehmen)")
                        else:
                            avg_row[col] = ''

                    # Füge die Durchschnitts-Zeile zum DataFrame hinzu
                    df = pd.concat([df, pd.DataFrame([avg_row])], ignore_index=True)

                    print(f"   ✅ Sub-Industry-Durchschnitt für '{sub_industry}' berechnet und hinzugefügt")
                else:
                    print(f"   ⚠️ Keine gültigen Daten für Sub-Industry: {sub_industry}")
            else:
                print(f"   ⚠️ Zu wenige Unternehmen für Sub-Industry: {sub_industry} (gefunden: {len(all_companies_in_sub_industry)})")

    # 2. FOCUS DURCHSCHNITTE (NUR FÜR FOCUS-GRUPPEN)
    print("   🎯 Berechne Focus Durchschnitte (nur für Gruppen mit Fokus)...")

    # Filtere nur die Zeilen mit vorhandenen Fokus-Werten
    df_focus_groups = df_numeric[df_numeric['Focus'].notna()]

    if not df_focus_groups.empty:
        # Hole alle eindeutigen Fokus-Werte
        unique_focus_values = df_focus_groups['Focus'].dropna().unique()

        for focus in unique_focus_values:
            if focus and focus.strip():
                print(f"     🔍 Suche alle Unternehmen mit Fokus: '{focus}'")

                # Hole ALLE Unternehmen mit diesem Fokus aus den Excel-Dateien
                all_companies_with_focus = find_companies_by_focus(focus)

                if len(all_companies_with_focus) > 1:
                    # Sammle Excel-Kennzahlen für ALLE Unternehmen mit diesem Fokus
                    all_focus_data = []
                    print(f"       �� Verarbeite {len(all_companies_with_focus)} Unternehmen...")

                    for i, company in enumerate(all_companies_with_focus, 1):
                        if i <= 5 or i % 20 == 0:  # Zeige nur jeden 20. nach den ersten 5
                            print(f"         {i}/{len(all_companies_with_focus)}: {company['Name']}")

                        company_data = get_kennzahlen_for_company(company['RIC'], excel_columns)
                        if company_data:
                            # Füge Basis-Informationen hinzu
                            company_data.update({
                                'Name': company['Name'],
                                'RIC': company['RIC'],
                                'Sub-Industry': company.get('Sub-Industry', ''),
                                'Focus': company.get('Focus', '')
                            })
                            all_focus_data.append(company_data)

                    if all_focus_data:
                        # Erstelle DataFrame für alle Unternehmen mit diesem Fokus
                        df_focus = pd.DataFrame(all_focus_data)

                        # Konvertiere zu numerischen Werten
                        for col in excel_columns:
                            df_focus[col] = pd.to_numeric(df_focus[col], errors='coerce')

                        # Berechne Durchschnitte
                        avg_row = {
                            'Name': f'🎯 Ø {focus}',
                            'RIC': '',
                            'GICS Sector': '',  # Hinzufügen der GICS Sector Spalte
                            'Sub-Industry': '',
                            'Focus': focus,
                            'Peer_Group_Type': 'Focus-Durchschnitt',
                            'Input_Row': f"Focus-Ø ({len(all_focus_data)} Unternehmen)",
                            'Input_Source': 'Durchschnitt (Fokus)'
                        }

                        for col in excel_columns:
                            valid_values = df_focus[col].dropna()
                            if len(valid_values) > 0:
                                avg_row[col] = valid_values.mean()
                                print(f"       📈 {col}: {avg_row[col]:.4f} (aus {len(valid_values)} Werten)")
                            else:
                                avg_row[col] = ''

                        # Füge die Durchschnitts-Zeile zum DataFrame hinzu
                        df = pd.concat([df, pd.DataFrame([avg_row])], ignore_index=True)

                        print(f"   ✅ Focus-Durchschnitt für '{focus}' berechnet und hinzugefügt")
                    else:
                        print(f"   ⚠️ Keine gültigen Daten für Fokus: {focus}")
                else:
                    print(f"   ⚠️ Zu wenige Unternehmen für Fokus: {focus} (gefunden: {len(all_companies_with_focus)})")
    else:
        print("   ⚠️ Keine Daten für Focus-Gruppen gefunden")

    return df

def calculate_refinitiv_averages_by_sector(df, refinitiv_fields):
    """Berechnet Sektor-Durchschnitte für Refinitiv-Kennzahlen basierend auf GICS-Sektoren - VEREINFACHT wie in der funktionierenden Version"""
    print("🔢 BERECHNE REFINITIV-DURCHSCHNITTE NACH SEKTOR...")

    if not refinitiv_fields or df.empty:
        print("⚠️ Keine Refinitiv-Kennzahlen oder leeres DataFrame, überspringe Durchschnittsberechnung")
        return df

    # Ermittle verwendete GICS Sectoren aus dem aktuellen DataFrame
    print("   🎯 Ermittle verwendete GICS-Sektoren aus verarbeiteten Unternehmen...")

    used_sectors = set()

    for _, row in df.iterrows():
        if not row['Name'].startswith('💼 Ø') and not row['Name'].startswith('🎯 Ø') and not row['Name'].startswith('🏭 Ø'):
            ric = row['RIC']
            if ric:
                sector = determine_gics_sector(ric)
                if sector:
                    used_sectors.add(sector)

    print(f"   📊 Verwendete GICS-Sektoren: {sorted(used_sectors)}")

    if not used_sectors:
        print("   ⚠️ Keine GICS-Sektoren identifiziert")
        return df

    # VEREINFACHTE LOGIK: Hole alle Sektor-Durchschnitte auf einmal
    print("   🌐 Hole Refinitiv-Sektor-Durchschnitte für alle verwendeten Sektoren...")
    all_sector_averages = get_all_sector_averages(list(used_sectors), refinitiv_fields)

    if not all_sector_averages:
        print("   ⚠️ Keine Sektor-Durchschnitte erhalten")
        return df

    # Erstelle Durchschnitts-Zeilen für jeden Sektor
    sector_average_rows = []

    for sector_name, sector_data in all_sector_averages.items():
        print(f"   ✅ Erstelle Sektor-Durchschnitts-Zeile für {sector_name}")

        # Erstelle Durchschnitts-Zeile
        avg_row = {
            'Name': f'🏭 Ø {sector_name}',
            'RIC': '',
            'GICS Sector': sector_name,
            'Sub-Industry': '',
            'Focus': '',
            'Peer_Group_Type': 'GICS-Sektor-Durchschnitt',
            'Input_Row': f'GICS-Sektor-Ø (Refinitiv-Branchendurchschnitt)',
        }

        # Füge alle Refinitiv-Kennzahlen hinzu
        for field in refinitiv_fields:
            clean_field = clean_refinitiv_field_name(field)

            # Suche nach dem Wert in den Sektor-Daten
            value = None
            for possible_key in [field, clean_field, field.replace('TR.', ''), clean_field]:
                if possible_key in sector_data:
                    value = sector_data[possible_key]
                    break

            if value is not None:
                avg_row[clean_field] = value
                print(f"       📊 {clean_field}: {value}")
            else:
                avg_row[clean_field] = ''

        sector_average_rows.append(avg_row)

    # Füge alle Sektor-Durchschnitte zum DataFrame hinzu
    if sector_average_rows:
        df_sectors = pd.DataFrame(sector_average_rows)
        df_combined = pd.concat([df, df_sectors], ignore_index=True)
        print(f"   ✅ {len(sector_average_rows)} GICS-Sektor-Durchschnitte hinzugefügt")
        return df_combined
    else:
        print("   ⚠️ Keine Sektor-Durchschnitte erstellt")
        return df
def save_beautiful_output(df, output_path):
    """Speichert das DataFrame mit verbesserter Formatierung und dynamischen Headern"""
    print("🎨 ERSTELLE VERBESSERTE EXCEL-AUSGABE...")

    # Berechne wie viele echte Unternehmen (ohne Durchschnitte) vorhanden sind
    actual_companies = df[~df['Name'].str.contains('Ø', na=False)]
    actual_company_count = len(actual_companies)

    print(f"   📊 {actual_company_count} echte Unternehmen, {len(df) - actual_company_count} Durchschnitte")

    # SPALTEN-KATEGORIEN DEFINIEREN (AM ANFANG!)
    company_cols = ['Name', 'RIC']  # Unternehmensdaten
    category_cols = ['Sub-Industry', 'Focus', 'GICS\nSektor']  # Kategorien
    metric_cols = [col for col in df.columns if col not in company_cols + category_cols + ['Peer_Group_Type', 'Input_Row', 'Sector']]  # Kennzahlen

    # 1. DYNAMISCHE HEADER-NAMEN basierend auf Dateninhalt
    df_formatted = df.copy()

    # Verbessere Header-Namen für bessere Lesbarkeit
    column_mapping = {}

    for col in df_formatted.columns:
        new_col_name = col

        # Spezielle Formatierung für verschiedene Spalten-Typen
        if col == 'Input_Row':
            new_col_name = 'Input\nZeile'
        elif col == 'Peer_Group_Type':
            new_col_name = 'Gruppe\nTyp'
        elif col == 'Sub-Industry':
            new_col_name = 'Sub-\nIndustry'
        elif 'TR.' in col:
            # Bereinige Refinitiv-Feldnamen für Header
            clean_name = clean_refinitiv_field_name(col)
            # Füge Zeilenumbrüche für bessere Darstellung hinzu
            if '(' in clean_name:
                parts = clean_name.split('(')
                new_col_name = f"{parts[0].strip()}\n({parts[1]}"
            else:
                new_col_name = clean_name
        elif len(col) > 15:
            # Lange Spalten-Namen umbrechen
            words = col.split(' ')
            if len(words) > 1:
                mid_point = len(words) // 2
                new_col_name = ' '.join(words[:mid_point]) + '\n' + ' '.join(words[mid_point:])

        if new_col_name != col:
            column_mapping[col] = new_col_name

    # Benenne Spalten um
    if column_mapping:
        df_formatted = df_formatted.rename(columns=column_mapping)
        print(f"   ✅ {len(column_mapping)} Spalten-Header optimiert")

    # 2. VERBESSERTE SEKTOR-INFORMATION HINZUFÜGEN
    print("   🏭 Verbessere GICS-Sektor-Information...")

    # Sammle alle GICS Sectoren für jede Zeile
    sectors = []
    for _, row in df_formatted.iterrows():
        ric = row.get('RIC', '')
        sector = ''

        if ric and not row['Name'].startswith('🏭 Ø') and not row['Name'].startswith('💼 Ø') and not row['Name'].startswith('🎯 Ø'):
            # Ermittle Sektor für echte Unternehmen anhand der Excel-Dateien
            for file in os.listdir(DATA_DIR):
                if file.endswith(".xlsx") and not file.startswith("~$"):
                    # Bestimme Sektor aus Dateiname (KORRIGIERT: Vollständige Mapping)
                    if "Consumer" in file and "Basic" not in file:
                        potential_sector = "Consumer Discretionary"
                    elif "Basic" in file and "Consumer" in file:
                        potential_sector = "Consumer Staples"
                    elif "Health" in file:
                        potential_sector = "Health Care"
                    elif "IT" in file or "Technology" in file:
                        potential_sector = "Information Technology"
                    elif "Materials" in file:
                        potential_sector = "Materials"
                    elif "Housing" in file:
                        potential_sector = "Real Estate"
                    elif "Utilities" in file:
                        potential_sector = "Utilities"
                    else:
                        continue

                    # Prüfe ob RIC in dieser Datei ist
                    file_path = os.path.join(DATA_DIR, file)
                    try:
                        xls = pd.ExcelFile(file_path)
                        for sheet_name in xls.sheet_names:
                            if any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                                try:
                                    df_check = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=100)
                                    if len(df_check.columns) >= 5:
                                        ric_col = df_check.columns[4]
                                        if any(df_check[ric_col].astype(str).str.upper().str.strip() == ric.upper().strip()):
                                            sector = potential_sector
                                            break
                                except:
                                    continue
                        if sector:
                            break
                    except:
                        continue

        elif row['Name'].startswith('🏭 Ø'):
            # Sektor-Durchschnitt - verwende den Sektor-Namen aus dem Name-Feld
            sector_name = row['Name'].replace('🏭 Ø ', '')
            sector = sector_name
        elif row['Name'].startswith('💼 Ø') or row['Name'].startswith('🎯 Ø'):
            # Sub-Industry oder Focus-Durchschnitte - versuche Sektor zu ermitteln
            # Schaue nach dem dominanten Sektor in der aktuellen Analyse
            non_avg_rows = df_formatted[~df_formatted['Name'].str.contains('Ø', na=False)]
            if not non_avg_rows.empty:
                # Nimm den häufigsten Sektor
                sector_counts = {}
                for _, non_avg_row in non_avg_rows.iterrows():
                    non_avg_ric = non_avg_row.get('RIC', '')
                    if non_avg_ric:
                        # Ermittle Sektor für diesen RIC
                        for file in os.listdir(DATA_DIR):
                            if file.endswith(".xlsx") and not file.startswith("~$"):
                                if "Consumer" in file and "Basic" not in file:
                                    check_sector = "Consumer Discretionary"
                                elif "Basic" in file and "Consumer" in file:
                                    check_sector = "Consumer Staples"
                                elif "Health" in file:
                                    check_sector = "Health Care"
                                elif "IT" in file or "Technology" in file:
                                    check_sector = "Information Technology"
                                elif "Materials" in file:
                                    check_sector = "Materials"
                                elif "Housing" in file:
                                    check_sector = "Real Estate"
                                elif "Utilities" in file:
                                    check_sector = "Utilities"
                                else:
                                    continue

                                file_path = os.path.join(DATA_DIR, file)
                                try:
                                    xls = pd.ExcelFile(file_path)
                                    for sheet_name in xls.sheet_names:
                                        if any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                                            try:
                                                df_check = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=100)
                                                if len(df_check.columns) >= 5:
                                                    ric_col = df_check.columns[4]
                                                    if any(df_check[ric_col].astype(str).str.upper().str.strip() == non_avg_ric.upper().strip()):
                                                        sector_counts[check_sector] = sector_counts.get(check_sector, 0) + 1
                                                        break
                                            except:
                                                continue
                                    if check_sector in sector_counts:
                                        break
                                except:
                                    continue

                # Verwende den häufigsten Sektor
                if sector_counts:
                    sector = max(sector_counts, key=sector_counts.get)

        sectors.append(sector)

    # Überschreibe oder füge GICS Sektor-Spalte hinzu
    df_formatted['GICS\nSektor'] = sectors
    print(f"   ✅ GICS-Sektor-Information für {len([s for s in sectors if s])} Zeilen hinzugefügt")

    # 3. REORDER SPALTEN für bessere Übersicht
    base_columns = ['Name', 'RIC', 'GICS\nSektor', 'Sub-Industry', 'Focus', 'Peer_Group_Type', 'Input_Row']

    # Sammle alle anderen Spalten (Kennzahlen)
    other_columns = [col for col in df_formatted.columns if col not in base_columns]

    # Neue Spalten-Reihenfolge
    new_column_order = []
    for col in base_columns:
        if col in df_formatted.columns:
            new_column_order.append(col)
    new_column_order.extend(other_columns)

    df_formatted = df_formatted[new_column_order]
    print(f"   ✅ Spalten neu angeordnet: {len(new_column_order)} Spalten")

    # 4. ERSTELLE EXCEL MIT PROFESSIONELLER FORMATIERUNG
    print("   🎨 Erstelle Excel-Formatierung...")

    # Speichere DataFrame als Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_formatted.to_excel(writer, sheet_name='Financial Analysis', index=False)

    # Lade Workbook für erweiterte Formatierung
    wb = load_workbook(output_path)
    ws = wb['Financial Analysis']

    # 🎨 ERWEITERTE FORMATIERUNG

    # Farb-Schema
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    alternating_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
    white_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    average_fill = PatternFill(start_color="e8f4fd", end_color="e8f4fd", fill_type="solid")  # Hellblau für Durchschnitte

    # Schrift-Stile
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    company_font = Font(name="Calibri", size=10, bold=True, color="1f4e79")
    average_font = Font(name="Calibri", size=10, bold=True, color="0066cc")
    data_font = Font(name="Calibri", size=9, color="2f2f2f")

    # Alignment
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center")

    # Borders
    thin_border = Border(
        left=Side(style="thin", color="b0b0b0"),
        right=Side(style="thin", color="b0b0b0"),
        top=Side(style="thin", color="b0b0b0"),
        bottom=Side(style="thin", color="b0b0b0")
    )

    thick_border = Border(
        left=Side(style="medium", color="1f4e79"),
        right=Side(style="medium", color="1f4e79"),
        top=Side(style="medium", color="1f4e79"),
        bottom=Side(style="medium", color="1f4e79")
    )

    # 5. TITEL HINZUFÜGEN
    ws.insert_rows(1)

    # Ermittle dominanten Sektor für Titel
    sector_counts = df_formatted['GICS\nSektor'].value_counts()
    main_sector = sector_counts.index[0] if not sector_counts.empty and sector_counts.index[0] else "Multi-Sector"

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📊 FINANCIAL PEER ANALYSIS - {main_sector.upper()}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1f4e79")
    title_cell.alignment = center_alignment

    # Titel über alle Spalten mergen
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_formatted.columns))
    ws.row_dimensions[1].height = 30

    # 6. HEADER FORMATIERUNG (Zeile 2)
    max_lines = 1
    for col_num, cell in enumerate(ws[2], 1):
        col_name = df_formatted.columns[col_num - 1]
        line_count = col_name.count('\n') + 1
        max_lines = max(max_lines, line_count)

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thick_border

    # Dynamische Header-Höhe
    header_height = max_lines * 15 + 10
    ws.row_dimensions[2].height = max(30, header_height)

    # 7. DATENZEILEN FORMATIERUNG
    for row_num in range(3, len(df_formatted) + 3):
        row_data = df_formatted.iloc[row_num - 3]

        for col_num, cell in enumerate(ws[row_num], 1):
            col_name = df_formatted.columns[col_num - 1]

            # Basis-Formatierung
            cell.border = thin_border
            cell.font = data_font
            cell.fill = white_fill  # Nur weiße Hintergrundfarbe

            # Spezielle Formatierung je Spalten-Typ
            if col_name in company_cols:
                cell.font = company_font if col_name == 'Name' else Font(name="Calibri", size=10, bold=True, color="1f4e79")
                cell.alignment = left_alignment
            elif col_name in category_cols:
                cell.alignment = center_alignment
            elif col_name in metric_cols:
                cell.alignment = right_alignment

                # Formatiere Zahlen schön
                if cell.value and str(cell.value).replace('.', '').replace('-', '').isdigit():
                    try:
                        num_val = float(cell.value)
                        if abs(num_val) >= 1:
                            cell.number_format = '#,##0.00'  # Mit Tausender-Trennzeichen
                        else:
                            cell.number_format = '0.0000'    # Mehr Dezimalstellen für kleine Zahlen
                    except:
                        pass
            else:
                cell.alignment = left_alignment

        # Zeilenhöhe optimieren
        ws.row_dimensions[row_num].height = 20

    # 8. CONDITIONAL FORMATTING FÜR DURCHSCHNITTE
    print("  🎯 Füge Conditional Formatting für Durchschnitte hinzu...")
    for row_num in range(3, len(df_formatted) + 3):
        row_data = df_formatted.iloc[row_num - 3]

        # Nur für Durchschnitts-Zeilen
        if row_data['Name'].startswith('Ø'):
            for col_num, cell in enumerate(ws[row_num], 1):
                col_name = df_formatted.columns[col_num - 1]

                if col_name in metric_cols:
                    # Hellblauer Hintergrund für Durchschnitts-Zeilen
                    cell.fill = average_fill

                    # Fett und blaue Schrift für Durchschnitts-Zeilen
                    cell.font = Font(name="Calibri", size=10, bold=True, color="0066cc")

    # 9. FINANZIELLE KENNZAHLEN FORMATIEREN
    print("  🎯 Formatiere finanzielle Kennzahlen...")
    for row_num in range(3, len(df_formatted) + 3):
        row_data = df_formatted.iloc[row_num - 3]

        for col_num, cell in enumerate(ws[row_num], 1):
            col_name = df_formatted.columns[col_num - 1]

            if col_name in metric_cols:
                # Wende Farbskalen-Regel an (grün-rot)
                try:
                    cell_rule = ColorScaleRule(
                        start_type="min",
                        start_color="FF0000",  # Rot
                        end_type="max",
                        end_color="00FF00"     # Grün
                    )
                    ws.conditional_formatting.add(f"{cell.coordinate}:{cell.coordinate}", cell_rule)
                except Exception as e:
                    print(f"     ⚠️ Fehler bei Conditional Formatting für {cell.coordinate}: {e}")

    # 10. LETZTE HANDGRIFFE UND SPEICHERN
    print("  ��� Letzte Handgriffe...")
    # Setze Standard-Schriftart für das gesamte Dokument
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.font is None or not cell.font.bold:
                    cell.font = data_font

    # Speichere die finale Datei
    wb.save(output_path)

    print(f"   💾 Verbesserte Excel-Ausgabe gespeichert: {output_path}")

def determine_gics_sector(ric):
    """Bestimmt den GICS Sektor für einen RIC anhand der Excel-Dateien"""
    if not ric:
        return ""

    for file in os.listdir(DATA_DIR):
        if not file.endswith(".xlsx") or file.startswith("~$"):
            continue

        # Bestimme Sektor aus Dateiname
        sector = None
        if "Consumer" in file and "Basic" not in file:
            sector = "Consumer Discretionary"
        elif "Basic" in file and "Consumer" in file:
            sector = "Consumer Staples"
        elif "Health" in file:
            sector = "Health Care"
        elif "IT" in file or "Technology" in file:
            sector = "Information Technology"
        elif "Materials" in file:
            sector = "Materials"
        elif "Housing" in file:
            sector = "Real Estate"
        elif "Utilities" in file:
            sector = "Utilities"
        else:
            continue

        # Prüfe ob RIC in dieser Datei ist
        file_path = os.path.join(DATA_DIR, file)
        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                if any(pattern in sheet_name.lower() for pattern in ["equity", "key", "revenue", "profitability", "financial", "growth", "figures"]):
                    try:
                        df_check = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=100)
                        if len(df_check.columns) >= 5:
                            ric_col = df_check.columns[4]
                            if any(df_check[ric_col].astype(str).str.upper().str.strip() == ric.upper().strip()):
                                return sector
                    except:
                        continue
        except:
            continue

    return ""

def get_gics_sector_mapping():
    """Mapping von GICS Sektor-Namen zu Refinitiv GICS Sektor-Nummern"""
    return {
        "Consumer Discretionary": "25",
        "Consumer Staples": "30",
        "Health Care": "35",
        "Information Technology": "45",
        "Materials": "15",
        "Real Estate": "60",
        "Utilities": "55",
        "Energy": "10",
        "Financials": "40",
        "Industrials": "20",
        "Communication Services": "50"
    }

def fetch_refinitiv_sector_averages(sector_name, refinitiv_fields):
    """Hole Refinitiv-Durchschnittsdaten für einen ganzen GICS Sektor"""
    # Diese Funktion delegiert an die neue Implementierung in refinitiv_integration.py
    from refinitiv_integration import fetch_refinitiv_sector_averages as new_fetch_function
    return new_fetch_function(sector_name, refinitiv_fields)
